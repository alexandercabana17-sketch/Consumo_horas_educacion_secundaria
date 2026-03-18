"""
Generador de Reporte de Cursos
Genera un Excel que muestra qué cursos considera el modelo después de aplicar
las equivalencias: exclusiones (Educación Inicial), cursos compartidos LLYA↔MYC
y cursos propios de cada programa.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colores de encabezado por hoja
COLOR_EXCLUIDOS   = "FFCCCC"   # rojo claro
COLOR_COMPARTIDOS = "CCE5FF"   # azul claro
COLOR_LLYA        = "D4EDDA"   # verde claro
COLOR_MYC         = "FFF3CD"   # amarillo claro
COLOR_RESUMEN     = "E2E2E2"   # gris claro


class GeneradorReporteCursos:
    """
    Genera el reporte Excel de cursos considerados en el análisis.

    Requiere que el AnalizadorHorasAula haya ejecutado:
      - cargar_datos()
      - identificar_cursos_compartidos()
      - identificar_cursos_a_eliminar()

    No requiere que se hayan procesado los programas ni generado resúmenes.
    """

    def __init__(self, analizador, output_path):
        """
        Inicializa el generador.

        Parameters
        ----------
        analizador : AnalizadorHorasAula
            Instancia con datos y equivalencias ya cargados.
        output_path : str
            Ruta de salida del archivo Excel.
        """
        self.analizador   = analizador
        self.output_path  = output_path

        print("\n" + "=" * 80)
        print("GENERADOR DE REPORTE DE CURSOS")
        print("=" * 80)

    # ------------------------------------------------------------------
    # Helpers privados para obtener datos
    # ------------------------------------------------------------------

    def _codigos_compartidos(self, programa):
        """Retorna el conjunto de códigos de cursos compartidos del programa."""
        key = 'codigo_llya' if programa == 'LLYA' else 'codigo_myc'
        return {cc[key] for cc in self.analizador.cursos_compartidos}

    def _filas_excluidos(self, programa):
        """Retorna lista de dicts para la hoja de cursos excluidos."""
        malla   = self.analizador.mallas[programa]
        equiv   = self.analizador.equivalencias[programa]
        codigos = self.analizador.cursos_a_eliminar[programa]

        filas = []
        for codigo in codigos:
            fila_malla = malla[malla['CODIGO_CURSO'] == codigo]
            fila_equiv = equiv[equiv['CODIGO_CURSO'] == codigo]

            if fila_malla.empty:
                continue

            row_m = fila_malla.iloc[0]
            motivo = ''
            if not fila_equiv.empty:
                motivo = fila_equiv.iloc[0]['PROGRAMA_EQUIVALENTE']

            filas.append({
                'Programa'   : programa,
                'Código'     : codigo,
                'Curso'      : row_m['CURSO'],
                'Semestre'   : row_m['SEMESTRE'],
                'Motivo'     : f'Equivale a {motivo}' if motivo else 'Excluido',
            })
        return filas

    def _filas_compartidos(self):
        """Retorna lista de dicts para la hoja de cursos compartidos."""
        filas = []
        for cc in self.analizador.cursos_compartidos:
            malla_llya = self.analizador.mallas['LLYA']
            malla_myc  = self.analizador.mallas['MYC']

            fila_llya = malla_llya[malla_llya['CODIGO_CURSO'] == cc['codigo_llya']]
            fila_myc  = malla_myc[malla_myc['CODIGO_CURSO']  == cc['codigo_myc']]

            if fila_llya.empty or fila_myc.empty:
                continue

            row_llya = fila_llya.iloc[0]
            ambientes = self.analizador.mapear_tipo_ambiente(row_llya)

            for tipo_amb, horas in ambientes:
                filas.append({
                    'Curso'          : cc['nombre'],
                    'Código LLYA'    : cc['codigo_llya'],
                    'Semestre LLYA'  : cc['semestre_llya'],
                    'Código MYC'     : cc['codigo_myc'],
                    'Semestre MYC'   : cc['semestre_myc'],
                    'Tipo Ambiente'  : tipo_amb,
                    'Horas Semanales': horas,
                    'Nota'           : (
                        'Fase prematuro: solo LLYA activo → cuenta sin fusión. '
                        'Fase compartida: ambos activos → sección combinada LLYA+MYC. '
                        'Fase solo MYC: MYC activo → cuenta sin fusión.'
                    ),
                })
        return filas

    def _filas_propios(self, programa):
        """Retorna lista de dicts para los cursos propios del programa."""
        malla              = self.analizador.mallas[programa]
        codigos_excluidos  = set(self.analizador.cursos_a_eliminar[programa])
        codigos_compartidos = self._codigos_compartidos(programa)

        filas = []
        for _, row in malla.iterrows():
            codigo = row['CODIGO_CURSO']
            if codigo in codigos_excluidos or codigo in codigos_compartidos:
                continue

            ambientes = self.analizador.mapear_tipo_ambiente(row)
            for tipo_amb, horas in ambientes:
                filas.append({
                    'Código'         : codigo,
                    'Curso'          : row['CURSO'],
                    'Semestre'       : row['SEMESTRE'],
                    'Tipo Ambiente'  : tipo_amb,
                    'Horas Semanales': horas,
                })
        return filas

    # ------------------------------------------------------------------
    # Helpers de formato
    # ------------------------------------------------------------------

    @staticmethod
    def _aplicar_estilo_encabezado(ws, color_hex):
        """Aplica estilo a la primera fila de la hoja."""
        fill   = PatternFill(fill_type='solid', fgColor=color_hex)
        font   = Font(bold=True)
        border = Border(
            bottom=Side(style='thin'),
        )
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.border    = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    @staticmethod
    def _ajustar_columnas(ws, min_width=10, max_width=50):
        """Ajusta el ancho de columnas al contenido."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ------------------------------------------------------------------
    # Creación de hojas
    # ------------------------------------------------------------------

    def _crear_hoja_resumen(self, writer):
        print("  Generando hoja: Resumen...")
        mallas = self.analizador.mallas
        excluidos = self.analizador.cursos_a_eliminar
        compartidos = self.analizador.cursos_compartidos

        total_llya = len(mallas['LLYA'])
        total_myc  = len(mallas['MYC'])
        excl_llya  = len(excluidos['LLYA'])
        excl_myc   = len(excluidos['MYC'])
        n_comp     = len(compartidos)

        efectivos_llya = total_llya - excl_llya
        efectivos_myc  = total_myc  - excl_myc
        # Cursos compartidos se cuentan una sola vez (viven en LLYA tras fusión)
        total_efectivo = efectivos_llya + efectivos_myc - n_comp

        filas = [
            ['REPORTE DE CURSOS CONSIDERADOS EN EL ANÁLISIS', ''],
            ['', ''],
            ['MALLA CURRICULAR LLYA', ''],
            ['  Total cursos en malla'              , total_llya],
            ['  Cursos excluidos (→ Ed. Inicial)'   , excl_llya],
            ['  Cursos compartidos con MYC'         , n_comp],
            ['  Cursos propios LLYA (efectivos)'    , efectivos_llya - n_comp],
            ['', ''],
            ['MALLA CURRICULAR MYC', ''],
            ['  Total cursos en malla'              , total_myc],
            ['  Cursos excluidos (→ Ed. Inicial)'   , excl_myc],
            ['  Cursos compartidos con LLYA'        , n_comp],
            ['  Cursos propios MYC (efectivos)'     , efectivos_myc - n_comp],
            ['', ''],
            ['RESUMEN DE CURSOS A CONSIDERAR', ''],
            ['  Cursos compartidos (sección combinada LLYA+MYC)', n_comp],
            ['  Cursos propios LLYA'                , efectivos_llya - n_comp],
            ['  Cursos propios MYC'                 , efectivos_myc - n_comp],
            ['  TOTAL CURSOS EFECTIVOS'             , total_efectivo],
            ['', ''],
            ['REGLAS DE EQUIVALENCIAS APLICADAS', ''],
            ['  1. Cursos → Ed. Inicial'            , 'EXCLUIDOS del análisis'],
            ['  2. Cursos compartidos LLYA↔MYC'     , 'Una sola sección con estudiantes combinados'],
            ['  2a. Periodo prematuro (solo LLYA)'  , 'Se cuenta como curso propio de LLYA'],
            ['  2b. Periodo compartido (ambos)'     , 'Fusión: sección única LLYA+MYC'],
            ['  2c. Periodo solo MYC'               , 'Se cuenta como curso propio de MYC'],
            ['  3. Cursos → otras carreras'         , 'SE MANTIENEN en el análisis'],
        ]

        df = pd.DataFrame(filas, columns=['Concepto', 'Valor'])
        df.to_excel(writer, sheet_name='Resumen', index=False)
        print("    OK: Resumen")

    def _crear_hoja_excluidos(self, writer):
        print("  Generando hoja: Cursos Excluidos...")
        filas = []
        for prog in ['LLYA', 'MYC']:
            filas.extend(self._filas_excluidos(prog))

        if filas:
            df = pd.DataFrame(filas)
        else:
            df = pd.DataFrame(columns=['Programa', 'Código', 'Curso', 'Semestre', 'Motivo'])

        df.to_excel(writer, sheet_name='Cursos Excluidos', index=False)
        print(f"    OK: {len(filas)} cursos excluidos")

    def _crear_hoja_compartidos(self, writer):
        print("  Generando hoja: Cursos Compartidos (Equivalencias LLYA-MYC)...")
        filas = self._filas_compartidos()

        if filas:
            df = pd.DataFrame(filas)
        else:
            df = pd.DataFrame(columns=[
                'Curso', 'Código LLYA', 'Semestre LLYA',
                'Código MYC', 'Semestre MYC',
                'Tipo Ambiente', 'Horas Semanales', 'Nota',
            ])

        df.to_excel(writer, sheet_name='Cursos Compartidos', index=False)
        print(f"    OK: {len(self.analizador.cursos_compartidos)} cursos compartidos")

    def _crear_hoja_propios(self, writer, programa):
        nombre_hoja = f'Cursos {programa}'
        print(f"  Generando hoja: {nombre_hoja}...")
        filas = self._filas_propios(programa)

        if filas:
            df = pd.DataFrame(filas)
        else:
            df = pd.DataFrame(columns=[
                'Código', 'Curso', 'Semestre', 'Tipo Ambiente', 'Horas Semanales',
            ])

        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        print(f"    OK: {len(filas)} filas (cursos×ambientes)")

    # ------------------------------------------------------------------
    # Punto de entrada público
    # ------------------------------------------------------------------

    def generar(self):
        """Genera el archivo Excel con el reporte de cursos."""
        print(f"\nGenerando reporte de cursos en: {self.output_path}")

        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            self._crear_hoja_resumen(writer)
            self._crear_hoja_excluidos(writer)
            self._crear_hoja_compartidos(writer)
            self._crear_hoja_propios(writer, 'LLYA')
            self._crear_hoja_propios(writer, 'MYC')

        # Aplicar formato post-escritura
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)

        colores = {
            'Resumen'             : COLOR_RESUMEN,
            'Cursos Excluidos'    : COLOR_EXCLUIDOS,
            'Cursos Compartidos'  : COLOR_COMPARTIDOS,
            'Cursos LLYA'         : COLOR_LLYA,
            'Cursos MYC'          : COLOR_MYC,
        }
        for nombre_hoja, color in colores.items():
            if nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                self._aplicar_estilo_encabezado(ws, color)
                self._ajustar_columnas(ws)

        wb.save(self.output_path)

        print("\n" + "=" * 80)
        print("REPORTE DE CURSOS GENERADO EXITOSAMENTE")
        print("=" * 80)
        print(f"\nArchivo: {self.output_path}")
        print("\nHojas generadas:")
        print("  1. Resumen             — conteo y reglas aplicadas")
        print("  2. Cursos Excluidos    — cursos que van a Educacion Inicial")
        print("  3. Cursos Compartidos  — equivalencias LLYA<->MYC y sus fases")
        print("  4. Cursos LLYA         — cursos propios del programa LLYA")
        print("  5. Cursos MYC          — cursos propios del programa MYC")
        print("=" * 80)
