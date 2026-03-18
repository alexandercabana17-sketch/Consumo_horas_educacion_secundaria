"""
Generador de Excel - Consumo de Horas-Aula
Genera un archivo Excel con:
  - Hoja "Tabla Pivote": resumen de horas por ambiente y periodo (con incrementos).
  - Una hoja por periodo (ej. "2027-01"): detalle de cada curso que contribuye
    a las horas de ese periodo, para verificación.
"""

import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Paleta sobria en escala de grises
COLOR_HEADER    = "404040"   # gris oscuro (texto blanco)
COLOR_SUBHEADER = "808080"   # gris medio (texto blanco)
COLOR_ALT_ROW   = "F2F2F2"   # gris muy claro para filas alternas
COLOR_SUBTOTAL  = "D9D9D9"   # gris claro para subtotales/totales


class GeneradorExcel:
    """
    Genera el archivo Excel de consumo de horas-aula.
    """

    def __init__(self, json_path, output_path):
        self.json_path   = json_path
        self.output_path = output_path

        with open(json_path, 'r', encoding='utf-8') as f:
            self.datos = json.load(f)

        print("\n" + "=" * 80)
        print("GENERADOR DE EXCEL")
        print("=" * 80)

    # ------------------------------------------------------------------
    # Hoja: Tabla Pivote
    # ------------------------------------------------------------------

    def crear_hoja_tabla_pivote(self, writer):
        print("\n  Generando hoja: Tabla Pivote...")

        if 'detalle_ambientes_especificos' not in self.datos:
            print("  ADVERTENCIA: No hay detalle de ambientes en el JSON.")
            return

        periodos = []
        for p in self.datos['detalle_ambientes_especificos']:
            row = {'Periodo': p['periodo']}
            for ambiente, info in sorted(p['ambientes'].items()):
                row[ambiente] = info['horas_semanales']
            row['Total'] = sum(i['horas_semanales'] for i in p['ambientes'].values())
            periodos.append(row)

        df = pd.DataFrame(periodos)

        # Identificar columnas de ambiente (todo excepto Periodo y Total)
        cols_amb = [c for c in df.columns if c not in ('Periodo', 'Total')]

        # Calcular columnas de incremento
        for col in cols_amb + ['Total']:
            inc = []
            for i, val in enumerate(df[col]):
                inc.append(val if i == 0 else val - df[col].iloc[i - 1])
            df[f'{col}_Incremento'] = inc

        # Ordenar columnas: Periodo | Aula valor+inc | Labs valor+inc | Taller | Virtual | Total
        ordered = ['Periodo']
        if 'Aula' in cols_amb:
            ordered += ['Aula', 'Aula_Incremento']
        for lab in sorted(c for c in cols_amb if 'Laboratorio' in c):
            ordered += [lab, f'{lab}_Incremento']
        if 'Taller' in cols_amb:
            ordered += ['Taller', 'Taller_Incremento']
        if 'Virtual' in cols_amb:
            ordered += ['Virtual', 'Virtual_Incremento']
        for col in cols_amb:
            if col not in ('Aula', 'Taller', 'Virtual') and 'Laboratorio' not in col:
                if col not in ordered:
                    ordered += [col, f'{col}_Incremento']
        ordered += ['Total', 'Total_Incremento']

        df = df[ordered]
        df.to_excel(writer, sheet_name='Tabla Pivote', index=False)
        print("    OK: Tabla Pivote")

    # ------------------------------------------------------------------
    # Hojas de detalle por periodo
    # ------------------------------------------------------------------

    def _nombre_hoja(self, periodo):
        """Convierte '2027-01' → '2027-I', '2027-02' → '2027-II'."""
        anio, ciclo = periodo.split('-')
        return f"{anio}-{'I' if ciclo == '01' else 'II'}"

    def crear_hojas_detalle_periodos(self, writer):
        print("\n  Generando hojas de detalle por periodo...")

        if 'detalle_cursos_por_periodo' not in self.datos:
            print("  ADVERTENCIA: No hay detalle de cursos en el JSON. Ejecute el analisis completo.")
            return

        COLS = ['Prog', 'Sem', 'Codigo', 'Curso', 'Tipo Ambiente',
                'Estudiantes', 'Hrs/Sem', 'Secciones', 'Total Hrs']

        for p in self.datos['detalle_cursos_por_periodo']:
            periodo  = p['periodo']
            cursos   = p['cursos']
            nombre   = self._nombre_hoja(periodo)

            filas = []
            for c in cursos:
                filas.append({
                    'Prog'         : c['programa'],
                    'Sem'          : c['semestre'],
                    'Codigo'       : c['codigo_curso'],
                    'Curso'        : c['curso'],
                    'Tipo Ambiente': c['tipo_ambiente'],
                    'Estudiantes'  : c['estudiantes'],
                    'Hrs/Sem'      : c['horas_semanales'],
                    'Secciones'    : c['secciones'],
                    'Total Hrs'    : c['horas_totales'],
                })

            df = pd.DataFrame(filas, columns=COLS) if filas else pd.DataFrame(columns=COLS)

            # Subtotales por tipo de ambiente
            subtotales = []
            ambientes_unicos = df['Tipo Ambiente'].unique() if not df.empty else []
            for amb in ambientes_unicos:
                sub = df[df['Tipo Ambiente'] == amb]['Total Hrs'].sum()
                subtotales.append({'Prog': 'SUBTOTAL', 'Sem': '', 'Codigo': '',
                                   'Curso': amb, 'Tipo Ambiente': '',
                                   'Estudiantes': '', 'Hrs/Sem': '',
                                   'Secciones': '', 'Total Hrs': sub})
            subtotales.append({'Prog': 'TOTAL', 'Sem': '', 'Codigo': '',
                               'Curso': '', 'Tipo Ambiente': '',
                               'Estudiantes': '', 'Hrs/Sem': '',
                               'Secciones': '', 'Total Hrs': df['Total Hrs'].sum() if not df.empty else 0})

            df_sub = pd.DataFrame(subtotales, columns=COLS)
            df_final = pd.concat([df, df_sub], ignore_index=True)
            df_final.to_excel(writer, sheet_name=nombre, index=False)

        total_hojas = len(self.datos['detalle_cursos_por_periodo'])
        print(f"    OK: {total_hojas} hojas de periodo generadas")

    # ------------------------------------------------------------------
    # Aplicar formato post-escritura
    # ------------------------------------------------------------------

    def _autofit(self, ws, min_w=8, max_w=50):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=min_w)
            ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

    def _formato_tabla_pivote(self, ws):
        fill_hdr  = PatternFill('solid', fgColor=COLOR_HEADER)
        fill_alt  = PatternFill('solid', fgColor=COLOR_ALT_ROW)
        font_hdr  = Font(bold=True, color='FFFFFF', size=10)
        font_body = Font(size=10)
        thin      = Side(style='thin', color='BFBFBF')
        border    = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border    = border

        for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
            fill = fill_alt if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill      = fill
                cell.font      = font_body
                cell.alignment = Alignment(horizontal='center' if cell.column == 1 else 'right')

        self._autofit(ws, min_w=10, max_w=40)

    def _formato_hoja_periodo(self, ws):
        fill_hdr = PatternFill('solid', fgColor=COLOR_HEADER)
        fill_sub = PatternFill('solid', fgColor=COLOR_SUBTOTAL)
        fill_alt = PatternFill('solid', fgColor=COLOR_ALT_ROW)
        font_hdr = Font(bold=True, color='FFFFFF', size=10)
        font_sub = Font(bold=True, size=10)
        font_bod = Font(size=10)
        thin     = Side(style='thin', color='BFBFBF')
        border   = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border

        for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
            prog_val = str(row[0].value or '')
            if prog_val in ('SUBTOTAL', 'TOTAL'):
                fill = fill_sub
                font = font_sub
            elif i % 2 == 0:
                fill = fill_alt
                font = font_bod
            else:
                fill = PatternFill()
                font = font_bod

            for cell in row:
                cell.fill      = fill
                cell.font      = font
                cell.alignment = Alignment(horizontal='center' if cell.column <= 2 else 'right')

        self._autofit(ws, min_w=8, max_w=50)

    def _aplicar_formato(self):
        wb = load_workbook(self.output_path)

        if 'Tabla Pivote' in wb.sheetnames:
            self._formato_tabla_pivote(wb['Tabla Pivote'])

        for nombre in wb.sheetnames:
            if nombre != 'Tabla Pivote':
                self._formato_hoja_periodo(wb[nombre])

        wb.save(self.output_path)

    # ------------------------------------------------------------------
    # Punto de entrada
    # ------------------------------------------------------------------

    def generar(self):
        """Genera el archivo Excel."""
        print(f"\nArchivo de salida: {self.output_path}")

        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            self.crear_hoja_tabla_pivote(writer)
            self.crear_hojas_detalle_periodos(writer)

        self._aplicar_formato()

        total_hojas = 1 + len(self.datos.get('detalle_cursos_por_periodo', []))
        print(f"\n  Total hojas generadas: {total_hojas}")
        print(f"\n{'=' * 80}")
        print("EXCEL GENERADO EXITOSAMENTE")
        print(f"{'=' * 80}")
        print(f"  Archivo: {self.output_path}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Uso: python generador_excel.py <json_path> <output_path>")
        sys.exit(1)
    GeneradorExcel(sys.argv[1], sys.argv[2]).generar()
